"""
共享 Excel 导出工具
"""
import io
from datetime import date
from decimal import Decimal
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def _safe_str(value) -> str:
    """安全转字符串：None→空串, date→YYYY-MM-DD, Decimal→str"""
    if value is None:
        return ""
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, Decimal):
        return str(value)
    return str(value)


def create_excel_response(filename: str, headers: list[str], rows: list[list]) -> io.BytesIO:
    """
    生成 Excel 文件并返回 BytesIO。

    - filename: 不含扩展名，示例 "算力租赁_合同列表_2026-07-23"
    - headers: 表头列表
    - rows: 数据行列表，每行是 list
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "合同列表"

    # 表头样式
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 写表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 写数据行
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_safe_str(value))
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # 自动列宽（大致估算）
    for col_idx in range(1, len(headers) + 1):
        max_width = len(str(headers[col_idx - 1])) * 2  # 中文约 2 字符宽
        for row_idx in range(2, len(rows) + 2):
            cell_val = str(ws.cell(row=row_idx, column=col_idx).value or "")
            max_width = max(max_width, len(cell_val) * 1.2)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_width + 2, 50)

    # 冻结首行
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
